from __future__ import annotations

from pathlib import Path
import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook, load_workbook

from aidas.services.step4_compiler import (
    DEFAULT_OUTPUT_FILENAME,
    MeasureCompilation,
    compile_step4_results,
    find_results_file,
)


def _write_results(path: Path, first_value: int) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "in"
    worksheet.cell(1, 7).value = "Round"
    for offset in range(20):
        worksheet.cell(offset + 2, 7).value = first_value + offset
    workbook.save(path)
    workbook.close()


def _write_thickness(path: Path, elm_value: float, onl_value: float) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        handle.write("Distance\tWhole\tELM\tONL\n")
        for distance in range(-1150, 1150):
            handle.write(f"{distance}\t0\t{elm_value}\t{onl_value}\n")


def _write_fovea(path: Path, value: float) -> None:
    path.write_text(f"foveaLIGHT 1 2\nRPEtoOLM {value} {value}\n", encoding="utf-8")


def _build_subject(root: Path) -> Path:
    subject = root / "LE" / "7_0_0"
    nasal = subject / "nasal"
    temporal = subject / "temporal"
    nasal.mkdir(parents=True)
    temporal.mkdir(parents=True)

    _write_results(nasal / "rr_MCPAR.xlsx", 1)
    _write_results(temporal / "Results.xlsx", 101)
    _write_thickness(nasal / "_thickness_vs_distance_from_fovea_LIGHT.txt", 10.0, 20.0)
    _write_thickness(temporal / "_thickness_vs_distance_from_fovea_LIGHT.txt", 30.0, 40.0)
    _write_fovea(nasal / "_fovea_light_profiles_LIGHT.txt", 55.77)
    _write_fovea(temporal / "_fovea_light_profiles_LIGHT.txt", 66.88)
    return subject


class Step4CompilerTests(unittest.TestCase):
    def test_compile_matches_legacy_layout_and_reads_current_step4_results(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            subject = _build_subject(root)
            output = root / DEFAULT_OUTPUT_FILENAME
            log_messages = []
            progress_events = []

            result = compile_step4_results(
                root,
                output,
                include_fovea=True,
                log_callback=log_messages.append,
                progress_callback=lambda completed, total, message: progress_events.append(
                    (completed, total, message)
                ),
            )

            self.assertEqual(result.output_path, output)
            self.assertEqual(result.le_ids, ("007",))
            self.assertEqual(result.re_ids, ())
            self.assertEqual(
                result.measures,
                (
                    MeasureCompilation("rrMCP-AR", 1, 0),
                    MeasureCompilation("ELM-RPE", 1, 0),
                    MeasureCompilation("ONL", 1, 0),
                ),
            )
            self.assertTrue(any("LE subject(s) found: 007" in line for line in log_messages))
            self.assertTrue(log_messages[-1].endswith(str(output)))
            self.assertEqual(progress_events[0], (0, 1, "Scanning input folders..."))
            self.assertEqual(progress_events[1][:2], (1, 5))
            self.assertEqual(
                [event[:2] for event in progress_events if event[2].startswith("Processed ")],
                [(2, 5), (3, 5), (4, 5)],
            )
            self.assertEqual(progress_events[-2], (4, 5, "Saving workbook..."))
            self.assertEqual(progress_events[-1], (5, 5, "Saved compiled workbook."))
            self.assertEqual(
                [completed for completed, _total, _message in progress_events],
                sorted(completed for completed, _total, _message in progress_events),
            )

            workbook = load_workbook(output, data_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["rrMCP-AR", "ELM-RPE", "ONL", "Extraction_log"],
                )

                rr_sheet = workbook["rrMCP-AR"]
                self.assertEqual(rr_sheet.max_row, 22)
                self.assertEqual(rr_sheet["A1"].value, "LE")
                self.assertEqual(rr_sheet["E1"].value, "RE")
                self.assertEqual(
                    [rr_sheet.cell(2, column).value for column in range(1, 8)],
                    ["ID", "Nasal", "Temporal", None, "ID", "Nasal", "Temporal"],
                )
                self.assertEqual(rr_sheet["A3"].value, "007")
                self.assertEqual(rr_sheet["B3"].value, 1)
                self.assertEqual(rr_sheet["B22"].value, 20)
                self.assertEqual(rr_sheet["C3"].value, 101)
                self.assertEqual(rr_sheet["C22"].value, 120)
                self.assertTrue(rr_sheet["A3"].font.bold)
                self.assertEqual(rr_sheet["A3"].alignment.horizontal, "center")
                self.assertEqual(rr_sheet["A3"].alignment.vertical, "center")
                self.assertEqual(rr_sheet["B3"].number_format, "0.000")
                for column in ("A", "B", "C", "E", "F", "G"):
                    self.assertEqual(rr_sheet.column_dimensions[column].width, 12.0)

                elm_sheet = workbook["ELM-RPE"]
                self.assertEqual(elm_sheet.max_row, 26)
                self.assertEqual(elm_sheet["A3"].value, "007")
                self.assertAlmostEqual(elm_sheet["B3"].value, 55.77)
                self.assertAlmostEqual(elm_sheet["C3"].value, 66.88)
                self.assertEqual(elm_sheet["B4"].value, 10.0)
                self.assertEqual(elm_sheet["B26"].value, 10.0)
                self.assertEqual(elm_sheet["C4"].value, 30.0)

                onl_sheet = workbook["ONL"]
                self.assertEqual(onl_sheet.max_row, 25)
                self.assertEqual(onl_sheet["A3"].value, "007")
                self.assertEqual(onl_sheet["B3"].value, 20.0)
                self.assertEqual(onl_sheet["B25"].value, 20.0)
                self.assertEqual(onl_sheet["C3"].value, 40.0)

                log_sheet = workbook["Extraction_log"]
                self.assertEqual(log_sheet.max_row, 4)
                self.assertEqual(log_sheet.max_column, 9)
                self.assertEqual(log_sheet["A2"].value, "rrMCP-AR")
                self.assertEqual(log_sheet["B2"].value, "LE")
                self.assertEqual(log_sheet["C2"].value, "007")
                self.assertEqual(log_sheet["I2"].value, "OK")
                self.assertTrue(str(log_sheet["E2"].value).endswith("rr_MCPAR.xlsx"))
                self.assertEqual(log_sheet["G2"].value, 20)
                self.assertEqual(log_sheet["G3"].value, 2300)
                self.assertTrue(log_sheet["A1"].font.bold)
                self.assertEqual(log_sheet["A1"].alignment.horizontal, "center")
                for column in "ABCDEFGHI":
                    self.assertEqual(log_sheet.column_dimensions[column].width, 28.0)
            finally:
                workbook.close()

            self.assertTrue(
                (subject / "nasal" / "_thickness_vs_distance_from_fovea_LIGHT2.txt").is_file()
            )
            self.assertTrue(
                (subject / "temporal" / "_thickness_vs_distance_from_fovea_LIGHT2.txt").is_file()
            )

    def test_fovea_can_be_excluded_from_elmrpe(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            _build_subject(root)
            output = root / "without_fovea.xlsx"

            compile_step4_results(root, output, include_fovea=False)

            workbook = load_workbook(output, data_only=True)
            try:
                elm_sheet = workbook["ELM-RPE"]
                self.assertEqual(elm_sheet.max_row, 25)
                self.assertEqual(elm_sheet["B3"].value, 10.0)
                self.assertEqual(elm_sheet["B25"].value, 10.0)
            finally:
                workbook.close()

    def test_current_step4_workbook_takes_precedence_over_legacy_results(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)
            _write_results(folder / "Results.xlsx", 100)
            current = folder / "rr_MCPAR.xlsx"
            _write_results(current, 1)

            self.assertEqual(Path(find_results_file(folder)), current)

    def test_skipped_subject_still_advances_each_measurement_progress_unit(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            nasal = root / "LE" / "8_0_0" / "nasal"
            nasal.mkdir(parents=True)
            workbook = Workbook()
            workbook.active.title = "in"
            workbook.active.cell(1, 7).value = "Round"
            workbook.save(nasal / "Results.xlsx")
            workbook.close()
            output = root / "skipped.xlsx"
            progress_events = []

            result = compile_step4_results(
                root,
                output,
                progress_callback=lambda completed, total, message: progress_events.append(
                    (completed, total, message)
                ),
            )

            self.assertEqual(
                result.measures,
                (
                    MeasureCompilation("rrMCP-AR", 0, 1),
                    MeasureCompilation("ELM-RPE", 0, 1),
                    MeasureCompilation("ONL", 0, 1),
                ),
            )
            self.assertEqual(
                [event[:2] for event in progress_events if event[2].startswith("Processed ")],
                [(2, 5), (3, 5), (4, 5)],
            )
            self.assertEqual(progress_events[-1][:2], (5, 5))

    def test_save_failure_does_not_report_full_progress(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            _build_subject(root)
            output = root / "locked.xlsx"
            progress_events = []

            with mock.patch(
                "aidas.services.step4_compiler.Workbook.save",
                side_effect=PermissionError("workbook is locked"),
            ):
                with self.assertRaisesRegex(PermissionError, "workbook is locked"):
                    compile_step4_results(
                        root,
                        output,
                        progress_callback=lambda completed, total, message: (
                            progress_events.append((completed, total, message))
                        ),
                    )

            self.assertEqual(progress_events[-1], (4, 5, "Saving workbook..."))
            self.assertFalse(
                any(completed == total for completed, total, _message in progress_events)
            )

    def test_empty_or_flat_folder_is_rejected_instead_of_saving_empty_output(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            _write_results(root / "Results.xlsx", 1)
            output = root / "empty.xlsx"

            with self.assertRaisesRegex(ValueError, "No LE or RE subject folders"):
                compile_step4_results(root, output)

            self.assertFalse(output.exists())


if __name__ == "__main__":
    unittest.main()
