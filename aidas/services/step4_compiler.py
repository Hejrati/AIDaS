"""Compile Step 4 measurements from an LE/RE folder tree into one workbook.

This module is the non-UI implementation of the legacy
``rrMCPAR_ELMRPE_ONL_Compiler`` utility.  It preserves the workbook layout,
row grouping, extraction order, and ``LIGHT2`` preparation behavior while
also recognizing the ``rr_MCPAR.xlsx`` files written by AIDaS Step 4.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
import glob
import os
from pathlib import Path
import re
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


ID_ROW_START = 3
RRMCPAR_BLOCK_SIZE = 20
ELMRPE_ONL_BLOCK_COUNT = 23
BIN_SIZE = 100
ELMRPE_COL = 3
ONL_COL = 4
FOVEA_LABEL = "RPEtoOLM"
SHEET_NAMES = ("rrMCP-AR", "ELM-RPE", "ONL")
DEFAULT_OUTPUT_FILENAME = "LE_RE_rrMCPAR_ELMRPE_ONL.xlsx"

PathLike = str | os.PathLike[str]
LogCallback = Callable[[str], None]
ProgressCallback = Callable[[int, int, str], None]
SubjectProcessedCallback = Callable[[str, str, str], None]


@dataclass(frozen=True)
class MeasureCompilation:
    """Included and skipped subject counts for one output sheet."""

    name: str
    included: int
    skipped: int


@dataclass(frozen=True)
class CompilationResult:
    """Summary returned after a compiler workbook is saved."""

    output_path: Path
    le_ids: tuple[str, ...]
    re_ids: tuple[str, ...]
    measures: tuple[MeasureCompilation, ...]


def classify_eye_from_path(path: PathLike) -> str:
    """Classify a subject folder as ``LE``, ``RE``, or ``UNKNOWN``."""

    parts = [part.upper() for part in os.path.normpath(os.fspath(path)).split(os.sep)]
    for part in parts:
        if "LE DONE" in part or part in ("LE", "LEFT") or part.startswith("LE "):
            return "LE"
        if "RE DONE" in part or part in ("RE", "RIGHT") or part.startswith("RE "):
            return "RE"

    joined = " | ".join(parts)
    if re.search(r"\bLE\b", joined):
        return "LE"
    if re.search(r"\bRE\b", joined):
        return "RE"
    return "UNKNOWN"


def find_subfolder_case_insensitive(parent_path: PathLike, target_name: str) -> str | None:
    parent = os.fspath(parent_path)
    if not os.path.isdir(parent):
        return None
    for entry in os.listdir(parent):
        full_path = os.path.join(parent, entry)
        if os.path.isdir(full_path) and entry.strip().lower() == target_name.lower():
            return full_path
    return None


def existing_file(*parts: PathLike) -> str | None:
    path = os.path.join(*(os.fspath(part) for part in parts))
    return path if os.path.isfile(path) else None


def find_results_file(folder_path: PathLike | None) -> str | None:
    """Find the current Step 4 workbook, then fall back to legacy names."""

    if not folder_path:
        return None
    folder = os.fspath(folder_path)
    if not os.path.isdir(folder):
        return None

    filenames = sorted(os.listdir(folder), key=str.lower)
    for filename in filenames:
        if filename.lower() == "rr_mcpar.xlsx" and not filename.startswith("~$"):
            path = os.path.join(folder, filename)
            if os.path.isfile(path):
                return path

    for filename in filenames:
        lower = filename.lower()
        if lower.startswith("results") and lower.endswith(".xlsx") and not filename.startswith("~$"):
            path = os.path.join(folder, filename)
            if os.path.isfile(path):
                return path
    return None


def find_nasal_temporal_results(subject_path: PathLike) -> tuple[str | None, str | None]:
    nasal_dir = find_subfolder_case_insensitive(subject_path, "nasal")
    temporal_dir = find_subfolder_case_insensitive(subject_path, "temporal")
    return find_results_file(nasal_dir), find_results_file(temporal_dir)


def get_first20_col_g_values(filepath: PathLike) -> list:
    workbook = load_workbook(os.fspath(filepath), data_only=True)
    try:
        worksheet = workbook.active
        values = []
        for row in worksheet.iter_rows(min_row=2, min_col=7, max_col=7, values_only=True):
            value = row[0]
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            values.append(value)
            if len(values) >= RRMCPAR_BLOCK_SIZE:
                break
        return values
    finally:
        workbook.close()


def extract_rrmcpar_from_subject(subject_path: PathLike):
    nasal_file, temporal_file = find_nasal_temporal_results(subject_path)
    nasal_values = get_first20_col_g_values(nasal_file) if nasal_file else []
    temporal_values = get_first20_col_g_values(temporal_file) if temporal_file else []
    return nasal_values, temporal_values, nasal_file, temporal_file


def parse_float_or_none(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() == "NA":
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def read_tsv_rows(filepath: PathLike) -> list[list[str]]:
    rows = []
    with open(filepath, "r", encoding="utf-8", errors="ignore", newline="") as handle:
        rows.extend(csv.reader(handle, delimiter="\t"))
    return rows


def write_tsv_rows(filepath: PathLike, rows: list[list[str]]) -> None:
    with open(filepath, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        writer.writerows(rows)


def block_means(
    values: list[float | None],
    block_count: int = ELMRPE_ONL_BLOCK_COUNT,
    bin_size: int = BIN_SIZE,
) -> list[float | None]:
    means = []
    for index in range(block_count):
        chunk = values[index * bin_size:(index + 1) * bin_size]
        numbers = [value for value in chunk if value is not None]
        means.append(sum(numbers) / len(numbers) if numbers else None)
    return means


def find_thickness_file(search_root: PathLike | None) -> str | None:
    if not search_root:
        return None
    root = os.fspath(search_root)
    if not os.path.isdir(root):
        return None
    patterns = (
        "**/*_thickness_vs_distance_from_fovea_LIGHT2*.txt",
        "**/*thickness*distance*fovea*LIGHT2*.txt",
        "**/*LIGHT2*.txt",
        "**/*light2*.txt",
        "**/*_thickness_vs_distance_from_fovea_LIGHT*.txt",
        "**/*thickness*distance*fovea*LIGHT*.txt",
        "**/*LIGHT*.txt",
        "**/*light*.txt",
    )
    for pattern in patterns:
        matches = sorted(glob.glob(os.path.join(root, pattern), recursive=True))
        matches = [match for match in matches if os.path.isfile(match)]
        if matches:
            return matches[0]
    return None


def find_light_files(subject_path: PathLike) -> tuple[str | None, str | None]:
    subject = os.fspath(subject_path)
    nasal_candidates = (
        existing_file(subject, "nasal", "_thickness_vs_distance_from_fovea_LIGHT2.txt"),
        existing_file(subject, "nasal", "_thickness_vs_distance_from_fovea_LIGHT.txt"),
        existing_file(subject, "rrMCPAR", "_thickness_vs_distance_from_fovea_LIGHT2.txt"),
        existing_file(subject, "rrMCPAR", "_thickness_vs_distance_from_fovea_LIGHT.txt"),
        existing_file(subject, "_thickness_vs_distance_from_fovea_LIGHT2.txt"),
        existing_file(subject, "_thickness_vs_distance_from_fovea_LIGHT.txt"),
    )
    temporal_candidates = (
        existing_file(subject, "temporal", "_thickness_vs_distance_from_fovea_LIGHT2.txt"),
        existing_file(subject, "temporal", "_thickness_vs_distance_from_fovea_LIGHT.txt"),
        existing_file(subject, "temporal", "rrMCPAR", "_thickness_vs_distance_from_fovea_LIGHT2.txt"),
        existing_file(subject, "temporal", "rrMCPAR", "_thickness_vs_distance_from_fovea_LIGHT.txt"),
    )

    nasal_file = next((path for path in nasal_candidates if path), None)
    temporal_file = next((path for path in temporal_candidates if path), None)

    if not nasal_file:
        for folder in ("nasal", "Nasal", "rrMCPAR"):
            nasal_file = find_thickness_file(os.path.join(subject, folder))
            if nasal_file:
                break
    if not temporal_file:
        for folder in ("temporal", "Temporal", "temp", "Temp", os.path.join("temporal", "rrMCPAR")):
            temporal_file = find_thickness_file(os.path.join(subject, folder))
            if temporal_file:
                break

    if not nasal_file or not temporal_file:
        all_text_files = []
        for root, _directories, filenames in os.walk(subject):
            for filename in filenames:
                lower = filename.lower()
                if (
                    lower.endswith(".txt")
                    and "thickness" in lower
                    and "distance" in lower
                    and "fovea" in lower
                    and "light" in lower
                ):
                    all_text_files.append(os.path.join(root, filename))
        all_text_files.sort(key=lambda path: ("light2" not in path.lower(), path.lower()))
        for filepath in all_text_files:
            lower_path = filepath.lower()
            if not temporal_file and (
                "temporal" in lower_path or "\\temp" in lower_path or "/temp" in lower_path
            ):
                temporal_file = filepath
            elif not nasal_file:
                nasal_file = filepath

    return nasal_file, temporal_file


def make_light2_from_light(light_path: PathLike | None, required_col: int) -> str | None:
    """Create the legacy cleaned ``LIGHT2`` file and return its path."""

    if not light_path:
        return None
    path = os.fspath(light_path)
    if "light2" in os.path.basename(path).lower():
        return path

    rows = read_tsv_rows(path)
    if not rows:
        return path

    cleaned = [rows[0]]
    kept_zero = False
    for row in rows[1:]:
        if len(row) < required_col:
            continue
        distance = parse_float_or_none(row[0])
        value = parse_float_or_none(row[required_col - 1])
        if distance == 0 and not kept_zero:
            cleaned.append(row)
            kept_zero = True
            continue
        if value is not None:
            cleaned.append(row)

    output_path = path[:-4] + "2.txt" if path.lower().endswith(".txt") else path + "2.txt"
    try:
        write_tsv_rows(output_path, cleaned)
        return output_path
    except PermissionError:
        return path


def get_column_values(light_path: PathLike | None, column_index: int) -> list[float]:
    if not light_path:
        return []
    path = os.fspath(light_path)
    if not os.path.isfile(path):
        return []
    values = []
    for row in read_tsv_rows(path)[1:]:
        if len(row) < column_index:
            continue
        value = parse_float_or_none(row[column_index - 1])
        if value is not None:
            values.append(value)
    return values


def find_fovea_file(folder_path: PathLike | None) -> str | None:
    if not folder_path:
        return None
    folder = os.fspath(folder_path)
    if not os.path.isdir(folder):
        return None
    candidates = (
        "_fovea_light_profiles_LIGHT.txt",
        "_fovea_light_profiles_LIGHT2.txt",
        "_fovea_light_LIGHT.txt",
        "_fovea_light_LIGHT2.txt",
        "fovea_light_profiles_LIGHT.txt",
        "fovea_light_profiles_LIGHT2.txt",
        "fovea_light_LIGHT.txt",
        "fovea_light_LIGHT2.txt",
    )
    for filename in candidates:
        path = existing_file(folder, filename)
        if path:
            return path
    for filename in os.listdir(folder):
        if "fovea_light" in filename.lower() and filename.lower().endswith(".txt"):
            path = os.path.join(folder, filename)
            if os.path.isfile(path):
                return path

    rr_folder = os.path.join(folder, "rrMCPAR")
    if os.path.isdir(rr_folder):
        for filename in candidates:
            path = existing_file(rr_folder, filename)
            if path:
                return path
        for filename in os.listdir(rr_folder):
            if "fovea_light" in filename.lower() and filename.lower().endswith(".txt"):
                path = os.path.join(rr_folder, filename)
                if os.path.isfile(path):
                    return path
    return None


def extract_fovea_value(filepath: PathLike | None, label: str = FOVEA_LABEL) -> float | None:
    if not filepath:
        return None
    path = os.fspath(filepath)
    if not os.path.isfile(path):
        return None
    with open(path, "r", encoding="utf-8", errors="ignore") as handle:
        for line in handle:
            parts = line.strip().split()
            if parts and parts[0].strip() == label and len(parts) >= 2:
                return parse_float_or_none(parts[1])
    return None


def row_count_for_measure(measure_name: str, include_fovea: bool) -> int:
    if measure_name == "rrMCP-AR":
        return RRMCPAR_BLOCK_SIZE
    if measure_name == "ELM-RPE":
        return ELMRPE_ONL_BLOCK_COUNT + 1 if include_fovea else ELMRPE_ONL_BLOCK_COUNT
    return ELMRPE_ONL_BLOCK_COUNT


def extract_elmrpe_onl_from_subject(
    subject_path: PathLike,
    measure_name: str,
    include_fovea: bool,
):
    column_index = ELMRPE_COL if measure_name == "ELM-RPE" else ONL_COL
    nasal_file, temporal_file = find_light_files(subject_path)
    nasal_light2 = make_light2_from_light(nasal_file, column_index) if nasal_file else None
    temporal_light2 = make_light2_from_light(temporal_file, column_index) if temporal_file else None

    nasal_values = get_column_values(nasal_light2, column_index) if nasal_light2 else []
    temporal_values = get_column_values(temporal_light2, column_index) if temporal_light2 else []
    nasal_blocks = block_means(nasal_values)
    temporal_blocks = block_means(temporal_values)

    if measure_name == "ELM-RPE" and include_fovea:
        subject = os.fspath(subject_path)
        nasal_fovea_file = find_fovea_file(os.path.join(subject, "nasal")) or find_fovea_file(subject)
        temporal_fovea_file = find_fovea_file(os.path.join(subject, "temporal"))
        nasal_blocks = [extract_fovea_value(nasal_fovea_file)] + nasal_blocks
        temporal_blocks = [extract_fovea_value(temporal_fovea_file)] + temporal_blocks

    row_count = row_count_for_measure(measure_name, include_fovea)
    nasal_blocks = (nasal_blocks + [None] * row_count)[:row_count]
    temporal_blocks = (temporal_blocks + [None] * row_count)[:row_count]
    return (
        nasal_blocks,
        temporal_blocks,
        nasal_file,
        temporal_file,
        len(nasal_values),
        len(temporal_values),
    )


def find_subject_folders(root_folder: PathLike) -> dict[str, dict[str, str]]:
    """Return discovered subject folders grouped by eye and zero-padded ID."""

    subject_folders: dict[str, dict[str, str]] = {"LE": {}, "RE": {}, "UNKNOWN": {}}
    for directory, _subdirectories, _filenames in os.walk(os.fspath(root_folder)):
        match = re.match(r"^(\d{1,3})(?:_|\b)", os.path.basename(directory))
        if not match:
            continue
        subject_id = match.group(1).zfill(3)
        nasal_results, temporal_results = find_nasal_temporal_results(directory)
        nasal_light, temporal_light = find_light_files(directory)
        if not any((nasal_results, temporal_results, nasal_light, temporal_light)):
            continue
        eye = classify_eye_from_path(directory)
        if subject_id not in subject_folders.setdefault(eye, {}):
            subject_folders[eye][subject_id] = directory
    return subject_folders


def write_eye_block(
    worksheet,
    log_worksheet,
    measure_name: str,
    eye: str,
    start_column: int,
    subject_folders_for_eye: dict[str, str],
    include_fovea: bool,
    subject_processed_callback: SubjectProcessedCallback | None = None,
) -> tuple[int, int]:
    id_column = start_column
    nasal_column = start_column + 1
    temporal_column = start_column + 2

    worksheet.cell(1, id_column).value = eye
    worksheet.cell(1, id_column).font = Font(bold=True)
    worksheet.cell(2, id_column).value = "ID"
    worksheet.cell(2, nasal_column).value = "Nasal"
    worksheet.cell(2, temporal_column).value = "Temporal"
    for column in (id_column, nasal_column, temporal_column):
        worksheet.cell(2, column).font = Font(bold=True)
        worksheet.cell(2, column).alignment = Alignment(horizontal="center")

    row_count = row_count_for_measure(measure_name, include_fovea)
    row = ID_ROW_START
    included = 0
    skipped = 0
    for subject_id in sorted(subject_folders_for_eye):
        subject_path = subject_folders_for_eye[subject_id]
        try:
            if measure_name == "rrMCP-AR":
                nasal_values, temporal_values, nasal_file, temporal_file = extract_rrmcpar_from_subject(
                    subject_path
                )
                nasal_count, temporal_count = len(nasal_values), len(temporal_values)
            else:
                (
                    nasal_values,
                    temporal_values,
                    nasal_file,
                    temporal_file,
                    nasal_count,
                    temporal_count,
                ) = extract_elmrpe_onl_from_subject(subject_path, measure_name, include_fovea)
        except Exception as exc:
            skipped += 1
            log_worksheet.append(
                [measure_name, eye, subject_id, subject_path, "", "", 0, 0, f"Error: {exc}"]
            )
            if subject_processed_callback is not None:
                subject_processed_callback(measure_name, eye, subject_id)
            continue

        if not any(value is not None for value in nasal_values) and not any(
            value is not None for value in temporal_values
        ):
            skipped += 1
            log_worksheet.append(
                [
                    measure_name,
                    eye,
                    subject_id,
                    subject_path,
                    nasal_file or "",
                    temporal_file or "",
                    nasal_count,
                    temporal_count,
                    "Skipped - no values extracted",
                ]
            )
            if subject_processed_callback is not None:
                subject_processed_callback(measure_name, eye, subject_id)
            continue

        id_cell = worksheet.cell(row, id_column)
        id_cell.value = subject_id
        id_cell.font = Font(bold=True)
        id_cell.alignment = Alignment(horizontal="center", vertical="center")
        for index in range(row_count):
            output_row = row + index
            nasal_cell = worksheet.cell(output_row, nasal_column)
            temporal_cell = worksheet.cell(output_row, temporal_column)
            nasal_cell.value = nasal_values[index] if index < len(nasal_values) else None
            temporal_cell.value = temporal_values[index] if index < len(temporal_values) else None
            if isinstance(nasal_cell.value, (int, float)):
                nasal_cell.number_format = "0.000"
            if isinstance(temporal_cell.value, (int, float)):
                temporal_cell.number_format = "0.000"

        log_worksheet.append(
            [
                measure_name,
                eye,
                subject_id,
                subject_path,
                nasal_file or "",
                temporal_file or "",
                nasal_count,
                temporal_count,
                "OK",
            ]
        )
        included += 1
        row += row_count
        if subject_processed_callback is not None:
            subject_processed_callback(measure_name, eye, subject_id)

    for column in (id_column, nasal_column, temporal_column):
        worksheet.column_dimensions[get_column_letter(column)].width = 12
    return included, skipped


def write_measure_sheet(
    worksheet,
    log_worksheet,
    measure_name: str,
    subject_folders: dict[str, dict[str, str]],
    include_fovea: bool,
    subject_processed_callback: SubjectProcessedCallback | None = None,
) -> tuple[int, int]:
    le_included, le_skipped = write_eye_block(
        worksheet,
        log_worksheet,
        measure_name,
        "LE",
        1,
        subject_folders.get("LE", {}),
        include_fovea,
        subject_processed_callback,
    )
    re_included, re_skipped = write_eye_block(
        worksheet,
        log_worksheet,
        measure_name,
        "RE",
        5,
        subject_folders.get("RE", {}),
        include_fovea,
        subject_processed_callback,
    )
    return le_included + re_included, le_skipped + re_skipped


def compile_step4_results(
    root_folder: PathLike,
    output_path: PathLike,
    *,
    include_fovea: bool = True,
    log_callback: LogCallback | None = None,
    progress_callback: ProgressCallback | None = None,
) -> CompilationResult:
    """Scan ``root_folder`` and save the formatted combined workbook."""

    root = Path(root_folder).expanduser()
    output = Path(output_path).expanduser()
    if not root.is_dir():
        raise ValueError("Choose a valid parent folder containing LE and RE subject folders.")
    if output.suffix.lower() != ".xlsx":
        raise ValueError("The compiled output filename must end with .xlsx.")
    if not output.parent.is_dir():
        raise ValueError("The selected output folder does not exist.")

    def emit(message: str) -> None:
        if log_callback is not None:
            log_callback(message)

    def report_progress(completed: int, total: int, message: str) -> None:
        if progress_callback is not None:
            progress_callback(int(completed), int(total), message)

    emit(f"Root folder: {root}")
    emit(f"Include fovea in ELM-RPE: {bool(include_fovea)}\n")
    emit("Scanning for LE and RE subject folders...")
    report_progress(0, 1, "Scanning input folders...")
    subject_folders = find_subject_folders(root)
    le_ids = tuple(sorted(subject_folders.get("LE", {})))
    re_ids = tuple(sorted(subject_folders.get("RE", {})))
    emit(f"LE subject(s) found: {', '.join(le_ids) or '(none)'}")
    emit(f"RE subject(s) found: {', '.join(re_ids) or '(none)'}\n")
    if not le_ids and not re_ids:
        raise ValueError(
            "No LE or RE subject folders with Step 4 data were found. Select the parent "
            "folder that contains the LE and RE folders, subject ID folders, and nasal or "
            "temporal data folders."
        )

    total_subjects = len(le_ids) + len(re_ids)
    total_steps = 2 + (len(SHEET_NAMES) * total_subjects)
    completed_steps = 1
    report_progress(
        completed_steps,
        total_steps,
        f"Found {total_subjects} subject folder{'s' if total_subjects != 1 else ''}.",
    )

    def subject_processed(measure_name: str, eye: str, subject_id: str) -> None:
        nonlocal completed_steps
        completed_steps += 1
        report_progress(
            completed_steps,
            total_steps,
            f"Processed {measure_name}: {eye} {subject_id}.",
        )

    workbook = Workbook()
    measure_summaries = []
    try:
        workbook.remove(workbook.active)
        log_worksheet = workbook.create_sheet("Extraction_log")
        log_worksheet.append(
            [
                "Measure",
                "Eye",
                "ID",
                "Subject folder",
                "Nasal file used",
                "Temporal file used",
                "Nasal raw values read",
                "Temporal raw values read",
                "Status",
            ]
        )

        for measure_name in SHEET_NAMES:
            worksheet = workbook.create_sheet(measure_name)
            included, skipped = write_measure_sheet(
                worksheet,
                log_worksheet,
                measure_name,
                subject_folders,
                bool(include_fovea),
                subject_processed,
            )
            measure_summaries.append(MeasureCompilation(measure_name, included, skipped))
            emit(f"{measure_name}: included {included}, skipped {skipped}")

        workbook._sheets = [workbook[name] for name in SHEET_NAMES] + [log_worksheet]
        for cell in log_worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for column in range(1, log_worksheet.max_column + 1):
            log_worksheet.column_dimensions[get_column_letter(column)].width = 28

        report_progress(completed_steps, total_steps, "Saving workbook...")
        workbook.save(output)
    finally:
        workbook.close()

    report_progress(total_steps, total_steps, "Saved compiled workbook.")
    emit(f"\nSaved: {output}")
    return CompilationResult(
        output_path=output,
        le_ids=le_ids,
        re_ids=re_ids,
        measures=tuple(measure_summaries),
    )


__all__ = [
    "CompilationResult",
    "DEFAULT_OUTPUT_FILENAME",
    "MeasureCompilation",
    "block_means",
    "classify_eye_from_path",
    "compile_step4_results",
    "extract_fovea_value",
    "find_results_file",
    "find_subject_folders",
    "row_count_for_measure",
]
